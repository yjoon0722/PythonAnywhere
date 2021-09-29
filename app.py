from flask import Flask, jsonify, request
import logging
import requests
from flask_cors import CORS
from urllib import parse
import urllib
import pymysql
import json
import time
import cryptocode
import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.utils import get_column_letter
import datetime
import re
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

app = Flask(__name__)
cors = CORS(app, support_credentials=True)

# logger instance 생성
logger = logging.getLogger(__name__)
# handler 생성 (stream, file)
streamHandler = logging.StreamHandler()
fileHandler = logging.FileHandler('./server.log')
# logger instance에 handler 설정
logger.addHandler(streamHandler)
logger.addHandler(fileHandler)
# logger instance로 log 찍기
logger.setLevel(level=logging.DEBUG)

handler = logging.FileHandler('./server.log', mode='a')
handler.setLevel(logging.INFO)
app.logger.addHandler(handler)

HEADERS = {'X-NCP-APIGW-API-KEY-ID': 'API-KEY-ID', 'X-NCP-APIGW-API-KEY': 'API-KEY'}

headers = {'Content-Type': 'application/json'}


# TimeZone 설정
os.environ["TZ"] = "Asia/Seoul"
time.tzset()

todayDate = datetime.datetime.now().date()
todayMonth = datetime.datetime.now().month
todayDay = datetime.datetime.now().day
shipmentOrderPath = "/home/intosharp/ReceiveData/{}/0_Send/".format(str(todayDate).replace("-",""))
juntech = "준테크"
goryeoHanjin = "고려(한진)"
goryeoGunyoung = "고려(건영)"

@app.route('/payapp', methods=['POST'])
def payapp():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST' :
        content = request.get_json(silent=True)
        # SSONG TODO: content null check

        # * 판매자 회원 아이디
        # SSONG TODO: 운영아이디로 교체 'hantongcorp'
        userid = 'hantongcorp'

        # * 상품명
        goodname = parse.quote(content['goodname'])

        # * 결제요청 금액
        price = content['price']
        # SSONG TODO: int type check

        # * 수신 휴대폰번호
        recvphone = content['recvphone']
        # SSONG TODO: phonenumber type check

        # 주소요청 (1:요청, 0:요청안함)
        reqaddr = '0'
        # 결제요청 SMS 발송여부 (n: SMS발송안함)
        smsuse = 'y'

        if not (goodname and price and recvphone):
            return jsonify({'error': 'No grguments'}), 400

        url = f'https://api.payapp.kr/oapi/apiLoad.html?cmd=payrequest&userid={userid}&goodname={goodname}&price={price}&recvphone={recvphone}&reqaddr={reqaddr}&smsuse={smsuse}'
        logger.debug('===================================== \nurl: %s', url)

        response = requests.request('POST', url)
        logger.debug('status code: %s', response.status_code)
        logger.debug('body: %s', response.text)
        logger.debug('===================================== \nResponse: %s', response)
        return jsonify({'result_code':'0','result_data':response.text}), response.status_code

    return jsonify()

# DB에 저장하면서 사용
@app.route('/payapp_feedback', methods=['POST'])
def payapp_feedback():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST' :
        mul_no = request.form.get('mul_no') # 결제요청번호
        recvphoneData = request.form.get('recvphone') # 구매자 연락처
        recvphone = ""
        try:
            recvphone = recvphoneData[:3] + "-" + recvphoneData[3:7] + "-" + recvphoneData[7:11]
        except:
            recvphone = recvphoneData
        goodname = request.form.get('goodname') # 상품명
        try:
            price = "{:,}".format(int(request.form.get('price'))) # 결제금액
        except:
            price = request.form.get('price') # 결제금액
        pay_state = request.form.get('pay_state') # 결제상태
        reqdate = request.form.get('reqdate') # 결제요청 일시
        pay_date = request.form.get('pay_date') # 결제승인 일시
        canceldate = request.form.get('canceldate') # 취소일시
        cancelmemo = request.form.get('cancelmemo') # 취소메모
        logger.debug("mul_no = %s",mul_no)
        logger.debug("recvphone = %s",recvphone)
        logger.debug("goodname = %s",goodname)
        logger.debug("price = %s",price)
        logger.debug("pay_state = %s",pay_state)
        logger.debug("reqdate = %s",reqdate)
        logger.debug("pay_date = %s",pay_date)
        logger.debug("canceldate = %s",canceldate)
        logger.debug("cancelmemo = %s",cancelmemo)

        try:
            db = dbcon()
            cursor = db.cursor()
            # 연락처로 DB에서 업체명 검색
            cursor.execute("SELECT account_code,account_name from account WHERE account_contact_information2 = %s OR account_contact_information1 = %s",(f"{recvphone}",f"{recvphone}"))
            account_data = cursor.fetchall()
            account_code = ""
            account_name = ""
            if len(account_data) > 0 :
                account_code = account_data[0][0] # 업체코드
                account_name = account_data[0][1] # 업체명
            logger.debug("len(account_data) = %s",len(account_data))
            logger.debug("account_code = %s",account_code)
            logger.debug("account_name = %s",account_name)
            if pay_state == "1" :
                # DB에 결제정보 업데이트
                # 결제요청번호 / 업체코드 / 업체명 / 연락처 / 상품명 / 결제금액 / 결제상태 / 결제요청 일시 / 결제승인 일시 / 취소일시 / 취소메모
                cursor.execute("INSERT INTO pay_app VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(f"{mul_no}",f"{account_code}",f"{account_name}",f"{recvphone}",f"{goodname}",f"{price}",f"{pay_state}",f"{reqdate}",f"{pay_date}",f"{canceldate}",f"{cancelmemo}"))
                db.commit()
                logger.debug("insert data and commit done")
            elif pay_state == "4" :
                cursor.execute("SELECT pay_state from pay_app WHERE mul_no = %s",(f"{mul_no}"))
                pay_state_data = cursor.fetchall()
                db_pay_state = pay_state_data[0][0]
                logger.debug("db_pay_state = %s",db_pay_state)
                if db_pay_state == "1" :
                    cursor.execute("UPDATE pay_app SET pay_state = %s WHERE mul_no = %s",(f"{pay_state}",f"{mul_no}"))
                    db.commit()
                    logger.debug("update data and commit done")
                    url = "https://intosharp.pythonanywhere.com/send_sms"
                    body = json.dumps({"recipients":[{"phoneNumber":"01092773322"}],"contents":"결제요청번호 : {}\n업체명 : {}\n구매자번호 : {}\n상품 : {}\n결제일시 : {}\n결제금액 : {}원\n결제가 완료됐습니다.".format(mul_no,account_name,recvphone,goodname,pay_date,price)})
                    _response = requests.post(url,data = body,headers = headers)
                    logger.debug("status_text = %s",_response.text)
            else :
                cursor.execute("SELECT pay_state from pay_app WHERE mul_no = %s",(f"{mul_no}"))
                pay_state_data = cursor.fetchall()
                db_pay_state = pay_state_data[0][0]
                if pay_state == db_pay_state :
                    pass
                else :
                    cursor.execute("UPDATE pay_app SET pay_state = %s,canceldate = %s, cancelmemo = %s WHERE mul_no = %s",(f"{pay_state}",f"{canceldate}",f"{cancelmemo}",f"{mul_no}"))
                    db.commit()
                    logger.debug("update data and commit done")


        except Exception as e:
            logger.debug("Exception : %s",e)
        finally:
            db.close()

        now = time.localtime()
        logger.debug('%04d/%02d/%02d %02d:%02d:%02d', now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
        # 결제요청시 checkretry이 'y'인 경우 응답이 'SUCCESS'가 아니면 재호출 합니다. (총 10회 / checkretry 기본값 'y')
        return "SUCCESS"

    return jsonify()

'''
# DB 미사용
@app.route('/payapp_feedback', methods=['POST'])
def payapp_feedback():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST' :
        mul_no = request.form.get('mul_no') # 결제요청번호
        recvphoneData = request.form.get('recvphone') # 구매자 연락처
        recvphone = ""
        try:
            recvphone = recvphoneData[:3] + "-" + recvphoneData[3:7] + "-" + recvphoneData[7:11]
        except:
            recvphone = recvphoneData
        goodname = request.form.get('goodname') # 상품명
        try:
            price = "{:,}".format(int(request.form.get('price'))) # 결제금액
        except:
            price = request.form.get('price') # 결제금액
        pay_state = request.form.get('pay_state') # 결제상태
        pay_date = request.form.get('pay_date') # 결제승인 일시
        logger.debug("mul_no = %s",mul_no)
        logger.debug("recvphone = %s",recvphone)
        logger.debug("goodname = %s",goodname)
        logger.debug("price = %s",price)
        logger.debug("pay_state = %s",pay_state)
        logger.debug("pay_date = %s",pay_date)

        try:
            db = dbcon()
            cursor = db.cursor()
            # 연락처로 DB에서 업체명 검색
            cursor.execute("SELECT account_name from account WHERE account_contact_information2 = %s OR account_contact_information1 = %s",(f"{recvphone}",f"{recvphone}"))
            account_data = cursor.fetchall()
            account_name = ""
            if len(account_data) > 0 :
                account_name = account_data[0][0] # 업체명
            logger.debug("account_name = %s",account_name)
            if pay_state == "4" :
                url = "https://intosharp.pythonanywhere.com/send_sms"
                body = json.dumps({"recipients":[{"phoneNumber":"01092773322"}],"contents":"결제요청번호 : {}\n업체명 : {}\n구매자번호 : {}\n상품 : {}\n결제일시 : {}\n결제금액 : {}원\n결제가 완료됐습니다.".format(mul_no,account_name,recvphone,goodname,pay_date,price)})
                _response = requests.post(url,data = body,headers = headers)
                logger.debug("status_text = %s",_response.text)

        except Exception as e:
            logger.debug("Exception : %s",e)
        finally:
            db.close()

        now = time.localtime()
        logger.debug('%04d/%02d/%02d %02d:%02d:%02d', now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
        # 결제요청시 checkretry이 'y'인 경우 응답이 'SUCCESS'가 아니면 재호출 합니다. (총 10회 / checkretry 기본값 'y')
        return "SUCCESS"

    return jsonify()
'''


# 주소를 위경도값으로 리턴
@app.route('/geocode', methods=['POST'])
def geocode():
    logger.debug('===================================== \nRequest: %s', request)

    if request.method == 'POST':
        content = request.get_json(silent=True)

        address = content['address'].strip()
        logger.debug('address: %s', address)

        # 주소
        query = parse.quote(address)
        logger.debug('query: %s', query)

        if not query:
            return jsonify({'error': 'No grgument'}), 400

        # 주소검색
        response = naveropenapi_geocode(address)
        if response:
            return jsonify({'result_code': '0', 'result_data': response.text}), response.status_code

        # 지역검색
        road_address = getRoadAddress(address)
        if road_address is None:
            return jsonify({'error': 'No roadAddress'}), 400
        road_response = naveropenapi_geocode(road_address)
        if road_response is None:
            return jsonify({'error': 'No roadAddress'}), 400
        return jsonify({'result_code': '0', 'result_data': road_response.text}), road_response.status_code

    return jsonify()


# 주소검색
def naveropenapi_geocode(address):
    url = f'https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode?query={address}'
    response = requests.request('GET', url, headers=HEADERS)
    logger.debug('body: %s', response.text)
    data = response.json()
    if not data:
        return None
    if not (data['status'] == 'OK'):
        return None
    if data['meta']['totalCount'] > 0:
        return response
    return None


# 지역검색
def getRoadAddress(address):
    url = f'https://openapi.naver.com/v1/search/local?query={address}&display=1&start=1&sort=random'
    response = requests.request('GET', url, headers={'X-Naver-Client-Id': 'hEbGczNWXDCzuehAe6mP', 'X-Naver-Client-Secret': 'SnpyDpIEiH'})
    logger.debug('body: %s', response.text)
    road_address = None
    try:
        data = response.json()
        if data and data['total'] > 0:
            road_address = data['items'][0]['roadAddress']
            if not road_address:
                road_address = data['items'][0]['address']
    except Exception as e:
        logger.debug('Error: getRoadAddress ', e)
    return road_address


# 출발지와 도착지에 대한 거리 계산
@app.route('/direction', methods=['POST'])
def direction():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST':
        content = request.get_json(silent=True)

        # 출발지
        start = parse.quote(content['start'])

        # 목적지. 하나 이상의 목적지 정보를 전달할 수 있으며, 복수 개의 목적지를 입력할 때는 :로 연결합니다. 입력한 목적지 정보 중 가장 적은 비용으로 도달할 수 있는 목적지로 경로가 생성됩니다.
        # 목적지의 최대 개수는 10개이며, 최초의 최적 목적지 좌표를 기준으로 직선거리 3km 이내의 좌표만 유효하게 판정합니다.
        goal = parse.quote(content['goal'])

        # 탐색 옵션. 옵션은 최대 3개까지 동시에 요청할 수 있으며, 여러 옵션은 ‘:‘로 연결합니다. 기본값은 traoptimal입니다.
        option = parse.quote(content['option'])

        if not (start and goal):
            return jsonify({'error': 'No grguments'}), 400

        url = f'https://naveropenapi.apigw.ntruss.com/map-direction-15/v1/driving?start={start}&goal={goal}&option={option}'
        logger.debug('===================================== \nurl: %s', url)

        response = requests.request('GET', url, headers=HEADERS)
        logger.debug('status code: %s', response.status_code)
        #logger.debug('body: %s', response.text)
        #logger.debug('===================================== \nResponse: %s', response)
        return jsonify({'result_code': '0', 'result_data': response.text}), response.status_code

    return jsonify()

# sql connect
def dbcon():
    return pymysql.connect(host='intosharp.mysql.pythonanywhere-services.com',user='userid',password='password',db='dbname',charset='utf8')

# find sales_status in sql
@app.route('/sales_status', methods=['POST'])
def sales_status():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST':
        content = request.get_json(silent=True)

        name = urllib.parse.unquote(parse.quote(content['name']))

        if not (name):
            return jsonify({'error': 'No grguments'}), 400

        try:
            db = dbcon()
            cursor = db.cursor()
            cursor.execute("""SELECT account_code FROM account WHERE account_name LIKE %s or account_contact_information1 LIKE %s or account_contact_information2 LIKE %s""",(f'%{name}%',f'%{name}%',f'%{name}%'))
            account_code_list = cursor.fetchall()
            account_sales_list = []

            try:
                for account_code in account_code_list:
                    cursor.execute(f"""SELECT a.account_name,b.goods,b.sales_date,a.account_contact_information1,a.account_contact_information2 FROM account as a LEFT JOIN sales_status as b ON a.account_code = b.account_code where a.account_code = '{account_code[0]}' ORDER BY b.sales_date DESC""")
                    results = cursor.fetchall()
                    if(len(results) > 5):
                        results = results[:5]

                    #account_sales_list.append(result)
                    row_headers=[x[0] for x in cursor.description] #this will extract row headers
                    for result in results:
                        account_sales_list.append(dict(zip(row_headers,result)))

            except Exception as ex:
                print('에러',ex)

            return jsonify({'result_code': '0', 'result_count': len(account_sales_list), 'result_array': json.dumps(account_sales_list)})

        finally:
            db.close()

    return jsonify()

# find account in sql
@app.route('/search_account', methods=['POST'])
def search_account():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST':
        content = request.get_json(silent=True)

        name = urllib.parse.unquote(parse.quote(content['name']))

        if not (name):
            return jsonify({'error': 'No grguments'}), 400

        try:
            db = dbcon()
            cursor = db.cursor()
            cursor.execute("""SELECT * FROM account WHERE account_name LIKE %s or account_contact_information1 LIKE %s or account_contact_information2 LIKE %s""",(f'%{name}%',f'%{name}%',f'%{name}%'))
            account_list = cursor.fetchall()
            account_list_detail = []
            try:
                row_headers=[x[0] for x in cursor.description] #this will extract row headers
                for account in account_list:
                    account_list_detail.append(dict(zip(row_headers,account)))
            except Exception as ex:
                print('에러',ex)

            return jsonify({'result_code': '0', 'result_count': len(account_list_detail), 'result_array': json.dumps(account_list_detail)})

        finally:
            db.close()

    return jsonify()

# Zone값 가져오기
def get_ecount_zone():
    post_data = {"COM_CODE":"COM_CODE"}
    url = 'https://oapi.ecount.com/OAPI/V2/Zone'

    response = requests.request('POST',url,data = json.dumps(post_data),headers = headers)
    # logger.debug('body: %s', response.text)
    if response.status_code == 200 :
        data = response.json()
        ZONE = data['Data']['ZONE']
    else:
        try:
            db = dbcon()
            cursor = db.cursor()
            cursor.execute('SELECT * from enter_order_url')
            enter_order_url_data = cursor.fetchall()
            ZONE = enter_order_url_data[0][0]
        finally:
            db.close()

    return ZONE

# 로그인 SESSION_ID 받아오기
def get_ecount_SESSION_ID(ZONE):
    post_data = {
        "COM_CODE":"COM_CODE",
        "USER_ID":"USER_ID",
        "API_CERT_KEY":"API-KEY",
        "LAN_TYPE":"ko-KR",
        "ZONE":ZONE
    }
    url = f'https://oapi{ZONE}.ecount.com/OAPI/V2/OAPILogin'
    response = requests.request('POST', url,data = json.dumps(post_data),headers = headers)
    # logger.debug('body: %s', response.text)

    try:
        db = dbcon()
        cursor = db.cursor()
        if response.status_code == 200 :
            data = response.json()
            SESSION_ID = data['Data']['Datas']['SESSION_ID']

            cursor.execute('TRUNCATE enter_order_url')
            cursor.execute('INSERT INTO enter_order_url VALUES(%s,%s)',(f"{ZONE}",f"{SESSION_ID}"))
            db.commit()
            # logger.debug('===================================== \ngood: %s', 'insert data done')
        else:
            cursor.execute('SELECT * from enter_order_url')
            enter_order_url_data = cursor.fetchall()
            SESSION_ID = enter_order_url_data[0][1]
    finally:
        db.close()

    return SESSION_ID

def get_etner_order_url():
    ZONE = get_ecount_zone()
    SESSION_ID = get_ecount_SESSION_ID(ZONE)

    url = f'https://oapi{ZONE}.ecount.com/OAPI/V2/SaleOrder/SaveSaleOrder?SESSION_ID={SESSION_ID}'

    return url

# ECOUNTERP 주문서 입력
@app.route('/ecounterp_enter_order', methods=['POST'])
def ecounterp_enter_order():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST' :
        content = request.get_json(silent=True)
        # SSONG TODO: content null check

        url = get_etner_order_url()
        logger.debug('===================================== \nurl: %s', url)

        response = requests.request('POST', url, data = json.dumps(content), headers = headers)
        logger.debug('status code: %s', response.status_code)
        logger.debug('body: %s', response.text)
        logger.debug('===================================== \nResponse: %s', response)
        return jsonify({'result_code':'0','result_data':response.text}), response.status_code

    return jsonify()


# find current sealing_vinyl in sql
@app.route('/current_sealing_vinyl', methods=['POST'])
def test():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST':
        content = request.get_json(silent=True)

        name = urllib.parse.unquote(parse.quote(content['name']))

        if not (name):
            return jsonify({'error': 'No grguments'}), 400

        try:
            db = dbcon()
            cursor = db.cursor()
            cursor.execute(f"""SELECT * FROM sales_status where account_code = '{name}' ORDER BY sales_date DESC Limit 1""")
            account_list = cursor.fetchall()
            account_list_detail = []
            try:
                row_headers=[x[0] for x in cursor.description] #this will extract row headers
                for account in account_list:
                    account_list_detail.append(dict(zip(row_headers,account)))
            except Exception as ex:
                print('에러',ex)

            return jsonify({'result_code': '0', 'result_count': len(account_list_detail), 'result_array': json.dumps(account_list_detail)})

        finally:
            db.close()

    return jsonify()

@app.route('/ad', methods=['GET'])
def ad():
    #ad = {"adURL" : "https://smartstore.naver.com/hantongbox/"}
    ad = {"adURL" : "http://zangzip.com/"}
    return jsonify(ad)

# 알린다 문자 전송
@app.route('/send_sms', methods=['POST'])
def send_sms():
    logger.debug('===================================== \nRequest: %s', request)
    try:
        content = request.get_json(silent=True)
        content['sender'] = {'phoneNumber': '01092773322'}
        content['contentsType'] = 'mms'
        content['type'] = '1'
        content['tags'] = []

        response = requests.request('POST',
            "https://shopapi.allinda.co.kr/api/v1.2/allinda/messages",
            data = json.dumps(content),
            headers = {"Content-Type": "application/json"},
            cookies = {"sessionid_textory": "sessionid_textory"}
            )
        logger.debug('status code: %s', response.status_code)
        logger.debug('body: %s', response.text)
        return jsonify(response.json()), response.status_code

        # return jsonify({'code': 500, 'content': content})
    except Exception as ex:
        logger.debug('Exception: send_sms - ', ex)
    return jsonify({'code': 400})


# 알린다 문자 전송 - 전송한 사용자 로그 남기기
@app.route('/send_sms_userid/<user_id>', methods=['POST'])
def send_sms_userid(user_id):
    logger.debug('===================================== \nRequest: %s', request)
    try:
        now = time.localtime()
        logger.debug('%04d/%02d/%02d %02d:%02d:%02d send user_id: %s', now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec, user_id)

        content = request.get_json(silent=True)
        content['sender'] = {'phoneNumber': '01092773322'}
        content['contentsType'] = 'mms'
        content['type'] = '1'
        content['tags'] = []

        response = requests.request('POST',
            "https://shopapi.allinda.co.kr/api/v1.2/allinda/messages",
            data = json.dumps(content),
            headers = {"Content-Type": "application/json"},
            cookies = {"sessionid_textory": "sessionid_textory"}
            )
        logger.debug('status code: %s', response.status_code)
        logger.debug('body: %s', response.text)
        return jsonify(response.json()), response.status_code

        # return jsonify({'code': 500, 'content': content})
    except Exception as ex:
        logger.debug('Exception: send_sms - ', ex)
    return jsonify({'code': 400})

# 발주서 데이터 전송
@app.route('/send_shipment_order_data', methods=['POST'])
def send_shipment_order_data():
    logger.debug('===================================== \nRequest: %s', request)
    if request.method == 'POST' :
        try:
            contentDataList = request.get_json(silent=True)['item']
            warehouse = request.get_json(silent=True)['warehouse']
            juntechContentList = []
            goryeoHanjinContentList = []
            goryeoGunyoungContentList = []
            for contentData in contentDataList :
                content = contentData['itemData']
                p_remarks2 = content['p_remarks2'].replace(" ","")
                trackMemo = setTrackMemo(p_remarks2)
                content['track_memo'] = trackMemo
                content['doc_qty'] = int(content['doc_qty'])
                del(content['p_remarks2'])
                del(content['isCheck'])
                if "준테크" in p_remarks2:
                    juntechContentList.append(list(content.values()))
                elif "고려" in p_remarks2:
                    for i in range(content['doc_qty']):
                        content['doc_qty'] = 1
                        goryeoHanjinContentList.append(list(content.values()))
                elif "건영" in p_remarks2:
                    for i in range(content['doc_qty']):
                        content['doc_qty'] = 1
                        goryeoGunyoungContentList.append(list(content.values()))

            createFolder(shipmentOrderPath)
            if warehouse == juntech :
                createShipmentOrderJuntech(juntechContentList)
                sendMail(warehouse)
            elif warehouse == goryeoHanjin :
                createShipmentOrderGoryeoHanjin(goryeoHanjinContentList)
                sendMail(warehouse)
            elif warehouse == goryeoGunyoung :
                createShipmentOrderGoryeoGunyoung(goryeoGunyoungContentList)
                sendMail(warehouse)
            elif warehouse == "전체" :
                createShipmentOrderJuntech(juntechContentList)
                sendMail(juntech)

                createShipmentOrderGoryeoHanjin(goryeoHanjinContentList)
                sendMail(goryeoHanjin)

                createShipmentOrderGoryeoGunyoung(goryeoGunyoungContentList)
                sendMail(goryeoGunyoung)

            return "SUCCESS"

        except Exception as e:
            logger.debug("Exception = %s",e)
            return "FAIL"

    return jsonify()

def setTrackMemo(p_remarks2):
    trackMemo = ""

    if "직접수령" in p_remarks2:
      trackMemo = "직접수령"
    elif "용달신용" in p_remarks2:
      trackMemo = "용달신용"
    elif "용달착불" in p_remarks2:
      trackMemo = "용달착불"
    elif "정기화물착불" in p_remarks2:
      trackMemo = "정기화물착불"
    elif "정기화물선불" in p_remarks2:
      trackMemo = "정기화물선불"

    return trackMemo

def createShipmentOrder(contentList):
    #엑셀 생성
    wb = openpyxl.Workbook()
    # 시트 지정
    ws = wb.active
    # 시트 이름변경
    ws.title = "송장"
    # 첫 Row에 리스트 append
    ws.append(["주문번호","상품명(옵션포함)","주문상품수량","주문자이름","주문자전화","주문자핸드폰","수취인이름","수취인전화","수취인핸드폰","신)수취인우편번호","수취인주소","송장번호","배송메시지"])


    rowCount = 2
    # 다음 Row부터 리스트 append
    for content in contentList:
        ws.append(content)
        # 배송메세지가 빈값이 아닐경우 배경색 노란색으로 변경
        if content[len(content)-1] != "" :
            for i in range(len(content)):
                ws[str(chr(i + 65)) + str(rowCount)].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                #chr(65) = A
        rowCount += 1

    # 셀 너비 조절
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if korlen(str(cell.value)) > max_length:
                    max_length = korlen(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 1
        ws.column_dimensions[column].width = adjusted_width

    return wb

# 준테크
def createShipmentOrderJuntech(contentList):
    wb = createShipmentOrder(contentList)
    juntech_count = get_shipment_order_count()[0][1] + 1
    wb.save(shipmentOrderPath + "{}준테크발주서-{}.xlsx".format(str(todayDate).replace("-",""),juntech_count))

# 고려포장(한진)
def createShipmentOrderGoryeoHanjin(contentList):
    wb = createShipmentOrder(contentList)
    goryeo_count = get_shipment_order_count()[0][2] + 1
    wb.save(shipmentOrderPath + "{}고려포장발주서-{}.xlsx".format(str(todayDate).replace("-",""),goryeo_count))
# 고려포장(건영)
def createShipmentOrderGoryeoGunyoung(contentList):
    wb = createShipmentOrder(contentList)
    goryeo_count = get_shipment_order_count()[0][2] + 1
    wb.save(shipmentOrderPath + "{}고려포장발주서-{}(건영택배).xlsx".format(str(todayDate).replace("-",""),goryeo_count))

# 영어 1글자, 한글 2글자로 글자수 계산하기
def korlen(str):
    korP = re.compile('[\u3131-\u3163\uAC00-\uD7A3]+',re.U)
    temp = re.findall(korP, str)
    temp_len = 0
    for item in temp:
        temp_len = temp_len + len(item)
    return len(str) + temp_len

# 폴더생성
def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except Exception as e:
        print('Error: ',e)

# 메일 보내기
def sendMail(warehouse):

    juntech_count = get_shipment_order_count()[0][1] + 1
    goryeo_count = get_shipment_order_count()[0][2] + 1
    logger.debug("juntech_count = %s",juntech_count)
    logger.debug("goryeo_count = %s",goryeo_count)


    sendEmail = "" # 보내는 계정 이메일
    password = "" # 보내는 계정 비밀번호
    # 실사용시 아래부분 주석해제
    # if warehouse == juntech :
    #     recvEmail = "joontech2016@naver.com" # 받는사람 이메일 (준테크)
    # elif warehouse == goryeoHanjin or warehouse == goryeoGunyoung:
    #     recvEmail = "no2kp@koreapk.com" # 받는사람 이메일 (고려)

    recvEmail = "joon_0722@naver.com" # 받는사람 이메일 (테스트계정)



    smtpName = "smtp.naver.com" #smtp 서버 주소
    smtpPort = 587 #smtp 포트 번호

    #여러 MIME을 넣기위한 MIMEMultipart 객체 생성
    msg = MIMEMultipart()

    # TODO: n차 발주서입니다 << 어떻게 해결?
    if warehouse == juntech :
        msg['Subject'] ="한통도시락 {}월 {}일_{}차 발주서 입니다.".format(todayMonth,todayDay,juntech_count)
    if warehouse == goryeoHanjin :
        msg['Subject'] ="한통도시락 {}월 {}일_{}차 발주서 입니다.".format(todayMonth,todayDay,goryeo_count)
    elif warehouse == goryeoGunyoung :
        msg['Subject'] ="한통도시락 {}월 {}일_{}차 건영 발주서 입니다.".format(todayMonth,todayDay,goryeo_count)
    msg['From'] = sendEmail
    msg['To'] = recvEmail

    if warehouse == juntech :
        text = """
안녕하세요.\n
한통도시락입니다.\n
{}월 {}일_{}차 발주서 첨부드립니다.\n
감사합니다.\n\n\n\n
한통도시락[010-9277-3322, hantongbox@naver.com]
Tel : 02-381-6318, Fax : 031-629-6318, http://zangzip.com\n
본 이메일에는 법률상 공개가 금지되거나 공개하여서는 안되는 비밀정보가 포함되어 있을 수 있습니다. 본 이메일을 받으신 분이 해당 수신인이 아니거나 또는 본 이메일을 해당 수신인에게 전달할 책임이 있는 직원 또는 대리인이 아닌 경우 본 이메일을 전파 또는 배포하거나, 복사하는 것은 엄격히 금지되어 있습니다. 만일 본 이메일이 잘못 전송되었을 경우에는 즉시 발신인에게 알려주시고 귀하의 컴퓨터에서 본 이메일을 삭제하여 주시기 바랍니다.
This e-mail message may contain legally privileged and/or confidential information. if you are not the intended recipient(s), or theemployee or agent responsible for delivery of this message to the intendedrecipient(s), you are hereby no tified that any dissemination, distribution orcopying of this e-mail message is strictly prohibited. If you have receivedthis message in error, please immediately notify the sender and delete thise-mail message from your computer.
""".format(todayMonth,todayDay,juntech_count)
    elif warehouse == goryeoHanjin :
        text = """
안녕하세요\n
한통 도시락입니다.\n
{}월 {}일_{}차 발주서 입니다.\n
감사합니다 :)
""".format(todayMonth,todayDay,goryeo_count)
    elif warehouse == goryeoGunyoung :
        text = """
안녕하세요\n
한통 도시락입니다.\n
{}월 {}일_{}차 건영 발주서 입니다.\n
감사합니다 :)
""".format(todayMonth,todayDay,goryeo_count)

    contentPart = MIMEText(text) #MIMEText(text , _charset = "utf8")
    msg.attach(contentPart)

    #파일 추가
    # fileList = []
    if warehouse == juntech :
        etcFileName = "{}준테크발주서-{}.xlsx".format(str(todayDate).replace("-",""),juntech_count)
        update_juntech_count(juntech_count)
    elif warehouse == goryeoHanjin :
        etcFileName = "{}고려포장발주서-{}.xlsx".format(str(todayDate).replace("-",""),goryeo_count)
        update_goryeo_count(goryeo_count)
    elif warehouse == goryeoGunyoung:
        etcFileName = "{}고려포장발주서-{}(건영택배).xlsx".format(str(todayDate).replace("-",""),goryeo_count)
        update_goryeo_count(goryeo_count)

    with open(shipmentOrderPath + etcFileName, 'rb') as etcFD :
        etcPart = MIMEApplication(etcFD.read())
        #첨부파일의 정보를 헤더로 추가
        etcPart.add_header('Content-Disposition','attachment', filename=etcFileName)
        msg.attach(etcPart)

    try:
        # 업체에 메일 발송
        s=smtplib.SMTP(smtpName ,smtpPort) #메일 서버 연결
        s.starttls() #TLS 보안 처리
        s.login(sendEmail, password) #로그인
        s.sendmail(sendEmail, recvEmail, msg.as_string()) #메일 전송, 문자열로 변환하여 보냅니다.
        s.close() #smtp 서버 연결을 종료합니다.
    except Exception as e:
        logger.debug('exception sendmail for warehouse: %s',e)

    try:
        # 네이버메일 보낸메일함에 저장이 안되기 때문에 한통계정으로 한번 더 발송
        recvEmail = "hantongbox@naver.com"
        s=smtplib.SMTP(smtpName ,smtpPort) #메일 서버 연결
        s.starttls() #TLS 보안 처리
        s.login(sendEmail, password) #로그인
        s.sendmail(sendEmail, recvEmail, msg.as_string()) #메일 전송, 문자열로 변환하여 보냅니다.
        s.close() #smtp 서버 연결을 종료합니다.
    except Exception as e:
        logger.debug('exception sendmail for hantongbox: %s',e)

def get_shipment_order_count() :
    db = dbcon()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM shipment_order WHERE send_date = %s",(f'{todayDate}'))
    shipment_order = cursor.fetchall()
    if len(shipment_order) < 1:
        cursor.execute("INSERT INTO shipment_order VALUES(%s,%s,%s)",(f'{todayDate}','0','0'))
        db.commit()
        cursor.execute("SELECT * FROM shipment_order WHERE send_date = %s",(f'{todayDate}'))
        shipment_order = cursor.fetchall()
    db.close()

    return shipment_order

def update_juntech_count(count):
    db = dbcon()
    cursor = db.cursor()
    cursor.execute("UPDATE shipment_order SET juntech_count = %s WHERE send_date = %s",(f'{count}',f'{todayDate}'))
    db.commit()
    db.close()

def update_goryeo_count(count):
    db = dbcon()
    cursor = db.cursor()
    cursor.execute("UPDATE shipment_order SET goryeo_count = %s WHERE send_date = %s",(f'{count}',f'{todayDate}'))
    db.commit()
    db.close()

if __name__ == '__main__':
    app.run(debug=True)
