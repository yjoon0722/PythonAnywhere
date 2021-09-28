import os
import datetime
import time

from itertools import groupby
from operator import itemgetter
from collections import defaultdict

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from key_generator.key_generator import generate

from common import *
from invoice_data import InvoiceData

# 택배 현현황판에 엑셀 문서 업로드하는 모듈입니다.
# pip install firebase-admin
# pip install openpyxl
# pip install key-generator

# SSONG TEST: 다운로드 루트 폴더 지정
# root_path = "/Users/songkiwon/Downloads/"

root_path = "/home/intosharp/ReceiveData/"
today = datetime.datetime.now()
todayDate = today.date()
tasks_time = datetime.datetime(today.year, today.month, today.day, 16,40) # tasks 등록 시간(pythonanywhere기준 07:35)

# tasks 등록시간으로부터 5분내로 실행하는 경우 당일 파일 검색, 아닌경우는 메일 발송날짜 파일 검색 (tasks시간 변경시 tasks_time 변경필요)
if tasks_time <= datetime.datetime.now() < tasks_time + datetime.timedelta(minutes=5) :
    date = str(todayDate).replace("-","")
    sendFileDate = todayDate
else:
    # 월:0 / 화:1 / 수:2 / 목:3 / 금:4 / 토:5 / 일:6
    if todayDate.weekday() == 0 :
        # 월요일인경우 금요일에 발송한 파일 검색
        sendFileDate = todayDate - datetime.timedelta(days=3)
    else:
        # 전날 발송메일 검색
        sendFileDate = todayDate - datetime.timedelta(days=1)

    date = str(sendFileDate).replace("-","")

# SSONG TEST: 날짜 지정
# todayDate = datetime.datetime.now().date()
# date = "20200914"

# 로딩된 데이타 파싱
def get_invoice_dict_list(wb):
  try:
    invoice_list = []
    for sheet in wb.sheetnames:
      carrier_name = get_carrier_name(sheet)
      for row in wb[sheet].iter_rows(min_row=2, min_col=1):
        if type(row) is EmptyCell:
          continue
        if row is None:
          continue

        timestamp = time.time()
        key = generate(separator='', min_atom_len=5, max_atom_len=5, capital='mix', seed=timestamp)

        invoice_data = InvoiceData()
        invoice_data.key          = key.get_key()
        invoice_data.fileName			= date                  # 파일명
        invoice_data.accountName  = row[1].value          # 판매처명
        invoice_data.warehouse		= row[2].value          # 창고
        invoice_data.phoneNumber  = row[3].value          # 모바일
        invoice_data.itemName			= row[4].value          # 품목명
        trackId = row[6].value if row[6].value is not None else ''
        invoice_data.trackId			= trackId               # 일반송장 번호
        invoice_data.carrierURL		= row[12].value if row[12].value is not None else '' # 택배사 배송 조회 URL
        invoice_data.carrierName  = carrier_name          # 택배사 이름
        invoice_data.carrierId		= get_carrier_id(sheet) # 택배사 코드

        invoice_list.append(invoice_data)

    invoice_list_sorted = sorted(invoice_list)
    invoice_dict = defaultdict(list)
    for invoice in invoice_list_sorted:
      invoice_dict[invoice.accountName].append(invoice)
    return invoice_dict
  except Exception as ex:
    # bot.send_message(chat_id = chat_id, text="송장번호 엑셀파일 업로드중 오류가 발생했습니다\n{}.".format(ex))
    print("Exception - get_invoice_dict_list: ", ex)
  finally:
    wb.close()

# 송장 파일 업로드
print("======== Start: Upload Invoice Excel File")
if todayDate.weekday() == 5 or todayDate.weekday() == 6 :
    pass
else :
  try:

    # Firebase 연결
    # SSONG TEST: Key 파일 경로 수정
    # cred = credentials.Certificate('./serviceAccountKey.json')
    cred = credentials.Certificate('/home/intosharp/project/serviceAccountKey.json')
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    # 엑셀 파일 로딩
    full_file_path = root_path + '{}/{}_택배송장.xlsx'.format(date, date)
    wb = load_workbook(full_file_path, read_only=True)
    invoice_dicts = get_invoice_dict_list(wb)

    fileName = date

    view_data_id   = u"ViewData" # ViewData
    track_data_id  = u"TrackData" # TrackData

    # ViewData Doc
    docRef = db.collection(view_data_id).document(fileName)

    # 업로드된 파일이 있는지 체크
    if docRef.get().exists:
      transaction = db.batch()
      # TrackData 삭제
      docs = db.collection(track_data_id).where(u'fileName', u'==', fileName).stream()
      for doc in docs:
        transaction.delete(doc.reference)
      # ViewData 삭제
      transaction.delete(docRef)
      transaction.commit()

    batch = db.batch()

    keys = []
    for key in invoice_dicts.keys():
      for val in invoice_dicts[key]:
        keys.append(val.key)

        print('set: ' + val.key + "-" + val.trackId)

        # ViewData Doc Field 추가
        batch.set(docRef, {val.key: val.__dict__}, merge=True)

        if not val.trackId.isdigit():
          continue

        # TrackData Doc 생성
        trackDataDocRef = db.collection(track_data_id).document(fileName+"_"+val.key)
        batch.set(trackDataDocRef,
          {
            u'key':        val.key,
            u'fileName':   val.fileName,
            u'trackId':    val.trackId,
            u'carrierURL': val.carrierURL,
            u'carrierId':  val.carrierId,
            u'statusId':   val.statusId,
          }
        )

    # ViewData Doc keys 추가
    weekday = "("+week_days[sendFileDate.weekday()]+")"
    batch.set(docRef, {u'keys':keys, u'weekday': weekday}, merge=True)

    batch.commit()

  except Exception as ex:
    # bot.send_message(chat_id = chat_id, text="송장번호 엑셀파일 업로드중 오류가 발생했습니다\n{}.".format(ex))
    print("Exception: ", ex)
  finally:
    wb.close()

  print("======== End: Upload Invoice Excel File")