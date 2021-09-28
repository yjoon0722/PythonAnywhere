import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

from order_data import OrderData

import datetime

from common import *

# 택배 현현황판에 올릴 엑셀 문서 생성하는 모듈입니다.
# pip install openpyxl

# SSONG TEST: 다운로드 루트 폴더 지정
# root_path = "/Users/songkiwon/Downloads/"

root_path = "/home/intosharp/ReceiveData/"
today = datetime.datetime.now()
todayDate = today.date()
tasks_time = datetime.datetime(today.year, today.month, today.day, 16,35) # tasks 등록 시간(pythonanywhere기준 07:35)

# tasks 등록시간으로부터 5분내로 실행하는 경우 당일 파일 검색, 아닌경우는 메일 발송날짜 파일 검색 (tasks시간 변경시 tasks_time 변경필요)
if tasks_time <= datetime.datetime.now() < tasks_time + datetime.timedelta(minutes=5) :
    date = str(todayDate).replace("-","")
else:
    # 월:0 / 화:1 / 수:2 / 목:3 / 금:4 / 토:5 / 일:6
    if todayDate.weekday() == 0 :
        # 월요일인경우 금요일에 발송한 파일 검색
        sendFileDate = todayDate - datetime.timedelta(days=3)
    else:
        # 전날 발송메일 검색
        sendFileDate = todayDate - datetime.timedelta(days=1)

    date = str(sendFileDate).replace("-","")

# SSONG TEST: 날짜 지정
# todayDate = datetime.datetime.now().date()
# date = "20210914"


# 엑셀 Rows 읽기
def get_excel_rows(full_file_path):
    try:
        order_list = []
        wb = load_workbook(full_file_path, read_only=True)
        for row in wb[wb.sheetnames[0]].iter_rows(min_row=2):
            order_list.append(OrderData(row))
    except Exception as ex:
        print("Exception - get_excel_rows: ", ex)
    finally:
        wb.close()
    return order_list

# 준테크 발주 문서 읽기
def get_joontech_orders():
    joontech_orders = {}
    try:
        dirname = root_path + "{}/0_Send/".format(date)
        filenames = os.listdir(dirname)
        for filename in filenames:
            if filename.startswith("~$") or filename.endswith(".xlsx") == False or not "준테크" in filename:
                continue
            full_filename = os.path.join(dirname, filename)
            orders = get_excel_rows(full_filename)
            for order in orders:
                joontech_orders[order.recipient_name] = order
    except Exception as ex:
        print("Exception - get_joontech_orders: ", ex)
    return joontech_orders

# 준테크 송장 문서 읽기
def get_joontech_rows(full_file_path, joontech_orders):
    try:
        order_list = []
        if not joontech_orders:
            return order_list
        wb = load_workbook(full_file_path, read_only=True)
        for row in wb[wb.sheetnames[0]].iter_rows():
            if type(row[1]) is EmptyCell:
                continue
            invoice = row[1].value.replace("-","")
            if not invoice.isdigit():
                continue
            key = row[4].value
            order_data = OrderData()
            joontech_order = joontech_orders[key]
            if joontech_order:
                order_data.order_number             = joontech_order.order_number
                order_data.product_name             = joontech_order.product_name
                order_data.product_quantity         = joontech_order.product_quantity
                order_data.order_name               = joontech_order.order_name
                order_data.order_phone              = joontech_order.order_phone
                order_data.order_mobile_phone       = joontech_order.order_mobile_phone
                order_data.recipient_name           = joontech_order.recipient_name
                order_data.recipient_phone          = joontech_order.recipient_phone
                order_data.recipient_mobile_phone   = joontech_order.recipient_mobile_phone
                order_data.recipient_post_code      = joontech_order.recipient_post_code
                order_data.recipient_address        = joontech_order.recipient_address
                order_data.invoice_number           = invoice
                order_data.delivery_message         = joontech_order.delivery_message
                order_list.append(order_data)
    except Exception as ex:
        print("Exception - get_joontech_rows: ", ex)
    finally:
        wb.close()
    return order_list

# 받은 송장 파일 리스트
print("======== Start: Create Invoice Excel File")
if todayDate.weekday() == 5 or todayDate.weekday() == 6 :
    pass
else :
    try:
        sheet_list = {}

        # 준테크 발주 목록
        joontech_orders = get_joontech_orders()

        # 엑셀 파일 로딩
        dirname = root_path + "{}/1_Receive/".format(date)
        filenames = os.listdir(dirname)
        for filename in filenames:
            if filename.startswith("~$") or filename.endswith(".xlsx") == False:
                continue

            full_filename = os.path.join(dirname, filename)

            if "고려포장" in filename:
                sheet_name = "고려포장(건영)" if "건영" in filename else "고려포장(한진)"
                if sheet_list.get(sheet_name):
                    sheet_list[sheet_name].extend(get_excel_rows(full_filename))
                else:
                    sheet_list[sheet_name] = get_excel_rows(full_filename)

            elif "한통 송장번호" in filename:
                sheet_name = "준테크(CJ)"
                if sheet_list.get(sheet_name):
                    sheet_list[sheet_name].extend(get_joontech_rows(full_filename, joontech_orders))
                else:
                    sheet_list[sheet_name] = get_joontech_rows(full_filename, joontech_orders)

        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active

        for key in sheet_list.keys():
            print("sheet name : ", key , len(sheet_list[key]))

            # Sheet 생성
            if ws is None:
                ws = wb.create_sheet(key)
            else:
                ws.title = key

            ColumnDimension(ws, bestFit=True)

            # Header 추가
            ws['A1'] = ''
            ws['B1'] = '판매처명'
            ws['C1'] = '창고'
            ws['D1'] = '모바일'
            ws['E1'] = '품목명'
            ws['F1'] = '정규식송장'
            ws['G1'] = '일반송장'
            ws['H1'] = '문구1'
            ws['I1'] = '문구2'
            ws['J1'] = '문구3'
            ws['K1'] = '기본URL'
            ws['L1'] = '주소복사대상'
            ws['M1'] = '배송조회URL'
            ws['N1'] = get_carrier_name(key)

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 10
            ws.column_dimensions['M'].width = 80
            ws.column_dimensions['N'].width = 20

            for idx, row in enumerate(sheet_list[key]):
                i = str(2 + idx)
                ws['A'+i] = idx + 1
                ws['B'+i] = row.recipient_name
                ws['C'+i] = get_warehouse_name(key)
                ws['D'+i] = row.recipient_mobile_phone
                ws['E'+i] = row.product_name
                ws['F'+i] = ''
                ws['G'+i] = get_invoice_number(row.invoice_number, row.delivery_message)
                ws['H'+i] = ''
                ws['I'+i] = ''
                ws['J'+i] = ''
                ws['K'+i] = ''
                ws['L'+i] = ''
                ws['M'+i] = get_carrier_url(key, row.invoice_number)
                ws['N'+i] = get_invoice_number(row.invoice_number, row.delivery_message)

            ws = None

        # wb.save(root_path + '{}/{}_택배송장_{}.xlsx'.format(date, date, str(datetime.datetime.now())))
        wb.save(root_path + '{}/{}_택배송장.xlsx'.format(date, date))

    except Exception as ex:
        bot.send_message(chat_id = chat_id, text="송장번호 엑셀파일 생성중 오류가 발생했습니다\n{}.".format(ex))
        print("Exception: ", ex)
    finally:
        wb.close()
    print("======== End: Create Invoice Excel File")
